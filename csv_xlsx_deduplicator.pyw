#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Скрипт для удаления дубликатов в CSV и XLSX файлах по первому столбцу (наименование).
- Оставляет только первую колонку.
- Убирает дубликаты по всей совокупности файлов.
- Разбивает результат на файлы по 3 млн строк (настраивается).
- Выходной формат: CSV с разделителем ';', заголовок 'Наименование', кодировка UTF-8-SIG (для Excel).
- Поддержка CSV в различных кодировках с автоматическим перебором и игнорированием ошибок.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import csv
import time
from datetime import datetime

# Попытка импорта openpyxl (для XLSX)
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

# ==================== НАСТРОЙКИ ПО УМОЛЧАНИЮ ====================
DEFAULT_MAX_LINES = 3_000_000          # 3 миллиона строк на файл
DEFAULT_SKIP_FIRST_ROW = False         # Пропускать ли первую строку во входных файлах (заголовок)

# Кодировки для CSV, которые будут перебираться (порядок важен)
CSV_ENCODINGS = ['utf-8-sig', 'cp1251', 'cp1252', 'latin-1', 'utf-8']


class CSVXLSXDeduplicatorApp:
    """Главный класс приложения."""

    def __init__(self, root):
        self.root = root
        self.root.title("Дедупликатор CSV/XLSX (по первому столбцу)")
        self.root.geometry("800x600")
        self.root.resizable(True, True)

        # Переменные состояния
        self.input_files = []            # Список путей к входным файлам
        self.is_processing = False
        self.stop_flag = False
        self.max_lines_per_file = DEFAULT_MAX_LINES
        self.skip_first_row = DEFAULT_SKIP_FIRST_ROW

        # Проверка наличия openpyxl
        if load_workbook is None:
            messagebox.showwarning(
                "Предупреждение",
                "Библиотека openpyxl не установлена. Файлы .xlsx обработаны не будут.\n"
                "Установите: pip install openpyxl"
            )

        self.create_widgets()

    def create_widgets(self):
        """Создание элементов интерфейса."""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # === Панель выбора файлов ===
        file_frame = ttk.LabelFrame(main_frame, text="Входные файлы", padding="5")
        file_frame.pack(fill=tk.X, pady=(0, 10))

        btn_frame = ttk.Frame(file_frame)
        btn_frame.pack(fill=tk.X, pady=5)

        self.btn_load_files = ttk.Button(btn_frame, text="📂 Выбрать файлы", command=self.load_files)
        self.btn_load_files.pack(side=tk.LEFT, padx=(0, 5))

        self.btn_load_folder = ttk.Button(btn_frame, text="📁 Выбрать папку", command=self.load_folder)
        self.btn_load_folder.pack(side=tk.LEFT, padx=(0, 5))

        self.btn_clear = ttk.Button(btn_frame, text="🗑️ Очистить список", command=self.clear_files)
        self.btn_clear.pack(side=tk.LEFT)

        # Список выбранных файлов
        list_frame = ttk.Frame(file_frame)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        scroll_y = ttk.Scrollbar(list_frame)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)

        self.files_listbox = tk.Listbox(
            list_frame,
            yscrollcommand=scroll_y.set,
            font=("Consolas", 9),
            height=6
        )
        self.files_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_y.config(command=self.files_listbox.yview)

        # Информация о файлах
        self.info_label = ttk.Label(file_frame, text="Файлов: 0 | Общий размер: 0 MB")
        self.info_label.pack(anchor=tk.W, pady=(5, 0))

        # === Настройки ===
        settings_frame = ttk.LabelFrame(main_frame, text="Настройки", padding="5")
        settings_frame.pack(fill=tk.X, pady=(0, 10))

        # Максимальное количество строк в выходном файле
        row_limit_frame = ttk.Frame(settings_frame)
        row_limit_frame.pack(fill=tk.X, pady=2)
        ttk.Label(row_limit_frame, text="Строк в файле:").pack(side=tk.LEFT)
        self.row_limit_var = tk.StringVar(value=str(DEFAULT_MAX_LINES))
        self.row_limit_entry = ttk.Entry(row_limit_frame, textvariable=self.row_limit_var, width=15)
        self.row_limit_entry.pack(side=tk.LEFT, padx=(5, 0))

        # Пропуск первой строки (заголовка) во входных файлах
        self.skip_first_var = tk.BooleanVar(value=DEFAULT_SKIP_FIRST_ROW)
        self.skip_first_check = ttk.Checkbutton(
            settings_frame,
            text="Пропускать первую строку во входных файлах (если есть заголовок)",
            variable=self.skip_first_var
        )
        self.skip_first_check.pack(anchor=tk.W, pady=2)

        # === Кнопки управления обработкой ===
        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill=tk.X, pady=(0, 10))

        self.btn_start = ttk.Button(control_frame, text="▶ Старт", command=self.start_processing, state=tk.DISABLED)
        self.btn_start.pack(side=tk.LEFT, padx=(0, 5))

        self.btn_stop = ttk.Button(control_frame, text="⏹ Стоп", command=self.stop_processing, state=tk.DISABLED)
        self.btn_stop.pack(side=tk.LEFT)

        # === Прогресс ===
        progress_frame = ttk.LabelFrame(main_frame, text="Прогресс", padding="5")
        progress_frame.pack(fill=tk.X, pady=(0, 10))

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill=tk.X, pady=(0, 5))

        self.status_label = ttk.Label(progress_frame, text="Ожидание...")
        self.status_label.pack(anchor=tk.W)

        self.stats_label = ttk.Label(progress_frame, text="")
        self.stats_label.pack(anchor=tk.W)

        # === Лог ===
        log_frame = ttk.LabelFrame(main_frame, text="Лог", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True)

        log_scroll = ttk.Scrollbar(log_frame)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self.log_text = tk.Text(
            log_frame,
            height=12,
            font=("Consolas", 9),
            yscrollcommand=log_scroll.set,
            state=tk.DISABLED,
            wrap=tk.WORD
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scroll.config(command=self.log_text.yview)

        # Теги для цветного лога
        self.log_text.tag_config("error", foreground="red")
        self.log_text.tag_config("success", foreground="green")
        self.log_text.tag_config("info", foreground="blue")
        self.log_text.tag_config("warn", foreground="orange")

    # ========== ВСПОМОГАТЕЛЬНЫЕ МЕТОДЫ ==========

    def log(self, message, tag=None):
        """Добавить сообщение в лог с временной меткой."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.config(state=tk.NORMAL)
        if tag:
            self.log_text.insert(tk.END, f"[{timestamp}] {message}\n", tag)
        else:
            self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)

    def update_file_info(self):
        """Обновить информацию о выбранных файлах (количество, общий размер)."""
        total_size = 0
        for f in self.input_files:
            try:
                total_size += os.path.getsize(f)
            except OSError:
                pass
        total_mb = total_size / (1024 * 1024)
        total_gb = total_size / (1024 * 1024 * 1024)
        if total_gb >= 1:
            size_str = f"{total_gb:.2f} GB"
        else:
            size_str = f"{total_mb:.2f} MB"
        self.info_label.config(text=f"Файлов: {len(self.input_files)} | Общий размер: {size_str}")

    def load_files(self):
        """Выбрать отдельные файлы через диалог."""
        if self.is_processing:
            messagebox.showwarning("Внимание", "Дождитесь завершения текущей операции")
            return
        files = filedialog.askopenfilenames(
            title="Выберите CSV или XLSX файлы",
            filetypes=[("Поддерживаемые файлы", "*.csv *.xlsx"), ("CSV", "*.csv"), ("XLSX", "*.xlsx"), ("Все файлы", "*.*")]
        )
        if files:
            self.input_files = list(files)
            self.input_files.sort()
            self.update_files_listbox()
            self.update_file_info()
            self.btn_start.config(state=tk.NORMAL)
            self.log(f"Загружено файлов: {len(self.input_files)}", "info")

    def load_folder(self):
        """Выбрать папку и добавить все CSV и XLSX файлы из неё (рекурсивно)."""
        if self.is_processing:
            messagebox.showwarning("Внимание", "Дождитесь завершения текущей операции")
            return
        folder = filedialog.askdirectory(title="Выберите папку с файлами")
        if folder:
            files = []
            for root_dir, _, filenames in os.walk(folder):
                for f in filenames:
                    if f.lower().endswith(('.csv', '.xlsx')):
                        files.append(os.path.join(root_dir, f))
            if files:
                self.input_files = files
                self.input_files.sort()
                self.update_files_listbox()
                self.update_file_info()
                self.btn_start.config(state=tk.NORMAL)
                self.log(f"Найдено файлов в папке: {len(self.input_files)}", "info")
            else:
                messagebox.showinfo("Информация", "В выбранной папке нет CSV или XLSX файлов.")

    def clear_files(self):
        """Очистить список файлов."""
        if self.is_processing:
            messagebox.showwarning("Внимание", "Дождитесь завершения текущей операции")
            return
        self.input_files = []
        self.files_listbox.delete(0, tk.END)
        self.update_file_info()
        self.btn_start.config(state=tk.DISABLED)
        self.log("Список файлов очищен", "info")

    def update_files_listbox(self):
        """Обновить отображение списка файлов."""
        self.files_listbox.delete(0, tk.END)
        for f in self.input_files:
            size_kb = os.path.getsize(f) / 1024
            name = os.path.basename(f)
            self.files_listbox.insert(tk.END, f"{name} ({size_kb:.1f} KB)")

    # ========== ОБРАБОТКА ==========

    def start_processing(self):
        """Запустить обработку в отдельном потоке."""
        if self.is_processing or not self.input_files:
            return

        # Считать настройки
        try:
            self.max_lines_per_file = int(self.row_limit_var.get())
            if self.max_lines_per_file <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Ошибка", "Количество строк в файле должно быть положительным целым числом.")
            return

        self.skip_first_row = self.skip_first_var.get()

        self.is_processing = True
        self.stop_flag = False

        # Блокировка кнопок
        self.btn_load_files.config(state=tk.DISABLED)
        self.btn_load_folder.config(state=tk.DISABLED)
        self.btn_clear.config(state=tk.DISABLED)
        self.btn_start.config(state=tk.DISABLED)
        self.btn_stop.config(state=tk.NORMAL)

        # Сброс прогресса
        self.progress_var.set(0)
        self.status_label.config(text="Подготовка...")
        self.stats_label.config(text="")

        # Запуск потока
        thread = threading.Thread(target=self.process_files, daemon=True)
        thread.start()

    def stop_processing(self):
        """Запросить остановку обработки."""
        self.stop_flag = True
        self.log("⏹ Получен сигнал остановки...", "warn")

    def process_files(self):
        """Основная функция обработки (выполняется в фоне)."""
        start_time = time.time()
        total_files = len(self.input_files)
        processed_files = 0
        total_rows_read = 0
        unique_rows = 0
        duplicate_rows = 0
        error_files = 0

        # Множество уникальных наименований
        seen_titles = set()

        # Переменные для текущего выходного файла
        output_dir = os.path.dirname(self.input_files[0])  # Берём директорию первого файла
        base_name = "deduplicated"
        part_num = 1
        lines_in_current_part = 0
        current_output = None

        def open_new_output():
            nonlocal current_output, part_num, lines_in_current_part
            if current_output:
                current_output.close()
            output_path = os.path.join(output_dir, f"{base_name}_part{part_num}.csv")
            # Записываем с BOM для совместимости с Excel
            current_output = open(output_path, 'w', encoding='utf-8-sig', newline='')
            writer = csv.writer(current_output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
            # Пишем заголовок
            writer.writerow(["Наименование"])
            self.log(f"📝 Создан выходной файл: {os.path.basename(output_path)}", "info")
            part_num += 1
            lines_in_current_part = 0

        # Открываем первый выходной файл
        open_new_output()

        try:
            for file_idx, file_path in enumerate(self.input_files):
                if self.stop_flag:
                    break

                file_name = os.path.basename(file_path)
                ext = os.path.splitext(file_name)[1].lower()
                self.root.after(0, lambda f=file_name: self.status_label.config(text=f"Обработка: {f}"))
                self.log(f"📄 Файл {file_idx+1}/{total_files}: {file_name}")

                file_rows = 0
                file_errors = 0

                try:
                    if ext == '.csv':
                        # Обработка CSV с перебором кодировок
                        file_obj = None
                        used_encoding = None
                        for enc in CSV_ENCODINGS:
                            try:
                                # Пробуем открыть и прочитать первую строку без игнорирования ошибок
                                f_test = open(file_path, 'r', encoding=enc, newline='')
                                f_test.readline()
                                f_test.close()
                                # Если успешно, открываем снова с errors='ignore' для всего файла
                                file_obj = open(file_path, 'r', encoding=enc, errors='ignore', newline='')
                                used_encoding = enc
                                break
                            except (UnicodeDecodeError, IOError):
                                continue
                        if file_obj is None:
                            raise Exception(f"Не удалось открыть файл ни в одной из кодировок: {CSV_ENCODINGS}")

                        self.log(f"   Используется кодировка: {used_encoding}")
                        reader = csv.reader(file_obj, delimiter=';')
                        for row_idx, row in enumerate(reader):
                            if self.stop_flag:
                                break
                            # Если нужно пропустить первую строку
                            if self.skip_first_row and row_idx == 0:
                                continue
                            # Проверяем, что строка не пуста и есть хотя бы одна колонка
                            if not row or len(row) == 0:
                                continue
                            title = row[0].strip()
                            if not title:
                                continue

                            total_rows_read += 1
                            file_rows += 1

                            if title in seen_titles:
                                duplicate_rows += 1
                            else:
                                seen_titles.add(title)
                                unique_rows += 1
                                writer = csv.writer(current_output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
                                writer.writerow([title])
                                lines_in_current_part += 1

                                # Проверка на необходимость нового файла
                                if lines_in_current_part >= self.max_lines_per_file:
                                    open_new_output()

                            # Обновление прогресса (каждые 1000 строк для плавности)
                            if total_rows_read % 1000 == 0:
                                self.update_progress(file_idx+1, total_files, total_rows_read, unique_rows, duplicate_rows)

                        file_obj.close()

                    elif ext == '.xlsx':
                        # Обработка XLSX (требуется openpyxl)
                        if load_workbook is None:
                            self.log(f"⚠️ Пропуск {file_name}: openpyxl не установлен", "error")
                            error_files += 1
                            continue

                        # Используем read_only=True для экономии памяти
                        wb = load_workbook(file_path, read_only=True)
                        # Берём активный лист (или первый)
                        sheet = wb.active
                        if sheet is None:
                            sheet = wb.worksheets[0]

                        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                            if self.stop_flag:
                                break
                            # Если нужно пропустить первую строку
                            if self.skip_first_row and row_idx == 0:
                                continue
                            if not row or len(row) == 0:
                                continue
                            cell_value = row[0]
                            if cell_value is None:
                                continue
                            title = str(cell_value).strip()
                            if not title:
                                continue

                            total_rows_read += 1
                            file_rows += 1

                            if title in seen_titles:
                                duplicate_rows += 1
                            else:
                                seen_titles.add(title)
                                unique_rows += 1
                                writer = csv.writer(current_output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
                                writer.writerow([title])
                                lines_in_current_part += 1

                                if lines_in_current_part >= self.max_lines_per_file:
                                    open_new_output()

                            if total_rows_read % 1000 == 0:
                                self.update_progress(file_idx+1, total_files, total_rows_read, unique_rows, duplicate_rows)

                        wb.close()  # Закрываем книгу

                except Exception as e:
                    self.log(f"❌ Ошибка при обработке {file_name}: {str(e)}", "error")
                    error_files += 1
                else:
                    self.log(f"   Строк в файле: {file_rows:,} (ошибок: {file_errors})")
                finally:
                    processed_files += 1
                    # Обновление прогресса после файла
                    self.update_progress(processed_files, total_files, total_rows_read, unique_rows, duplicate_rows)

        finally:
            # Закрыть последний выходной файл
            if current_output:
                current_output.close()
                # Если последний файл пустой (не содержит данных кроме заголовка) — удалить его
                if lines_in_current_part == 0:
                    try:
                        os.remove(current_output.name)
                        self.log(f"🗑️ Удалён пустой выходной файл: {os.path.basename(current_output.name)}")
                    except OSError:
                        pass

        elapsed = time.time() - start_time

        # Финальная статистика
        self.root.after(0, lambda: self.show_final_stats(
            total_files, processed_files, error_files,
            total_rows_read, unique_rows, duplicate_rows,
            part_num - 1, elapsed
        ))

    def update_progress(self, files_done, total_files, rows_read, unique, dup):
        """Обновить элементы прогресса (вызывается из фонового потока)."""
        def _update():
            if total_files > 0:
                progress = (files_done / total_files) * 100
                self.progress_var.set(progress)
            self.status_label.config(text=f"Обработано файлов: {files_done}/{total_files}")
            self.stats_label.config(
                text=f"Строк: {rows_read:,} | Уникальных: {unique:,} | Дублей: {dup:,}"
            )
        self.root.after(0, _update)

    def show_final_stats(self, total_files, processed, errors, total_rows, unique, dup, parts, elapsed):
        """Показать итоговую статистику в логе и диалоге."""
        self.log("=" * 60)
        self.log(f"✅ ОБРАБОТКА ЗАВЕРШЕНА за {elapsed:.1f} сек", "success")
        self.log(f"   Всего файлов: {total_files} (обработано: {processed}, ошибок: {errors})")
        self.log(f"   Прочитано строк: {total_rows:,}")
        self.log(f"   Уникальных записей: {unique:,}")
        self.log(f"   Дубликатов удалено: {dup:,}")
        self.log(f"   Создано выходных файлов: {parts}")
        self.log("=" * 60)

        messagebox.showinfo(
            "Готово",
            f"Обработка завершена!\n\n"
            f"Файлов обработано: {processed} / {total_files}\n"
            f"Ошибок: {errors}\n"
            f"Уникальных записей: {unique:,}\n"
            f"Дубликатов удалено: {dup:,}\n"
            f"Выходных файлов: {parts}\n"
            f"Время: {elapsed:.1f} сек"
        )

        # Разблокировка кнопок
        self.btn_load_files.config(state=tk.NORMAL)
        self.btn_load_folder.config(state=tk.NORMAL)
        self.btn_clear.config(state=tk.NORMAL)
        self.btn_start.config(state=tk.NORMAL)
        self.btn_stop.config(state=tk.DISABLED)
        self.is_processing = False
        self.status_label.config(text="Готово")


if __name__ == "__main__":
    root = tk.Tk()
    app = CSVXLSXDeduplicatorApp(root)
    root.mainloop()